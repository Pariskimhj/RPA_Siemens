# 엑셀 읽어오기
import datetime
import pyautogui
from openpyxl import load_workbook
import pyperclip
import webbrowser
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

webbrowser.open_new("https://unipass.customs.go.kr/csp/index.do")
for i in range(2, 6):
    i = str(i)
    wb = load_workbook('수입현황_'+i+'(22.12.21).xlsx') # 수입 파일 열기
    ws = wb.active
    col_B = ws["B"]
    # 전체 BL 조회하기
    ws['U1'] = '진행상태'
    ws['T1'] = '통관진행상태'
    count = 0
    for cell in col_B:
        # title 제외
        if cell.row == 1:
            continue
        if ws["L"+cell.coordinate[1:]].value < datetime.datetime.today(): # 입항일자가 오늘 이전이고
            if ws["O"+cell.coordinate[1:]].value == False or ws["P"+cell.coordinate[1:]].value == False: # 통관완료 또는 반출완료가 False이면
                count += 1
                # 화물진행정보 들어가기
                if count == 1:
                    pyautogui.click(-122, 1294, duration=4.5) # 화물진행정보 B/L 클릭 : 회사용
                    #pyautogui.click(1352, 317, duration=4.5)
                    pyautogui.click(-128, 1321) # B/L 넣는 칸 클릭 : 회사용
                    #pyautogui.click(1335, 364)

                    bl = cell.value
                    pyperclip.copy(bl)
                    pyautogui.hotkey('ctrl', 'v') # bl 붙여넣기
                    pyautogui.press('enter') # Enter

                    # 진행 상태 긁어오기
                    pyautogui.moveTo(-468,1484, 1.5) # 회사용
                    #pyautogui.moveTo(836,605, 1.5)
                    pyautogui.dragTo(-363,1520, 1) # 회사용
                    #pyautogui.dragTo(988,662, 1)
                    pyautogui.hotkey('ctrl', 'c') # 결과 복사
                    result = pyperclip.paste() # 결과 저장
                    ws["U"+cell.coordinate[1:]] = result

                    # 통관 진행 상태 긁어오기
                    pyautogui.moveTo(-674,1593, 1.5) # 회사용
                    pyautogui.dragTo(-576,1625, 1) # 회사용
                    pyautogui.hotkey('ctrl', 'c') # 결과 복사
                    result = pyperclip.paste() # 결과 저장
                    ws["T"+cell.coordinate[1:]] = result
                else:
                    # BL 붙여 넣기
                    bl = cell.value # BL number
                    pyperclip.copy(bl) # 클립보드에 bl 복사
                    pyautogui.click(-362, 1331, duration=1.5) # 회사용
                    #pyautogui.click(993, 372, duration=1.5)
                    pyautogui.hotkey('ctrl', 'v')
                    pyautogui.press('enter')

                    # 진행 상태 긁어오기
                    #pyautogui.moveTo(832, 608, 1.5)
                    pyautogui.moveTo(-468,1484, 1.5) # 회사용
                    #pyautogui.dragTo(989, 659, 1)
                    pyautogui.dragTo(-363,1520, 1) # 회사용
                    pyautogui.hotkey('ctrl', 'c') # 결과 복사
                    result = pyperclip.paste() # 결과 저장
                    ws["U"+cell.coordinate[1:]] = result

                    # 통관 진행 상태 긁어오기
                    pyautogui.moveTo(-674,1593, 1.5) # 회사용
                    pyautogui.dragTo(-576,1625, 1) # 회사용
                    pyautogui.hotkey('ctrl', 'c') # 결과 복사
                    result = pyperclip.paste() # 결과 저장
                    ws["T"+cell.coordinate[1:]] = result
        # 입항일자가 오늘 이후면 title 노란색
        else:
            ws["B"+cell.coordinate[1:]].fill = PatternFill(fgColor="ffff00", fill_type='solid')
    col_U = ws["U"]
    # 진행상태
    for cell in col_U:
        if cell.value == '수입신고 수리완료':
            if ws["O"+cell.coordinate[1:]].value == False:
                ws["O"+cell.coordinate[1:]].value = True
                ws["O"+cell.coordinate[1:]].fill = PatternFill(fgColor="ffff00", fill_type='solid') # 통관 완료 노란색
                ws["B"+cell.coordinate[1:]].fill = PatternFill(fgColor="0066ff", fill_type='solid') # title 파란색
        elif cell.value == '반출완료':
            #if ws["O"+cell.coordinate[1:]].value == False:
            #    ws["O"+cell.coordinate[1:]].value = True
            #    ws["O"+cell.coordinate[1:]].fill = PatternFill(fgColor="ffff00", fill_type='solid') # 통관 완료 노란색
            #    ws["B"+cell.coordinate[1:]].fill = PatternFill(fgColor="0066ff", fill_type='solid') # title 파란색
            if ws["T"+cell.coordinate[1:]].value == '수입(사용소비) 심사진행': # 보세건
                ws["B"+cell.coordinate[1:]].fill = PatternFill(fgColor="ff0000", fill_type='solid') # title 빨간색
                ws["P"+cell.coordinate[1:]].value = True
                ws["P"+cell.coordinate[1:]].fill = PatternFill(fgColor="ffff00", fill_type='solid') # 반출 완료 노란색
            elif ws["P"+cell.coordinate[1:]].value == False:
                ws["P"+cell.coordinate[1:]].value = True
                ws["P"+cell.coordinate[1:]].fill = PatternFill(fgColor="ffff00", fill_type='solid') # 반출 완료 노란색
                ws["B"+cell.coordinate[1:]].fill = PatternFill(fgColor="0066ff", fill_type='solid') # title 파란색

    col_T = ws["T"]
    for cell in col_T:
        if cell.value == '수입신고수리':
            if ws["O"+cell.coordinate[1:]].value == False:
                ws["O"+cell.coordinate[1:]].value = True
                ws["O"+cell.coordinate[1:]].fill = PatternFill(fgColor="ffff00", fill_type='solid') # 통관 완료 노란색
                ws["B"+cell.coordinate[1:]].fill = PatternFill(fgColor="0066ff", fill_type='solid') # title 파란색
    wb.save('수입현황_'+i+'(22.12.21).xlsx') # 파일 저장
    wb.close