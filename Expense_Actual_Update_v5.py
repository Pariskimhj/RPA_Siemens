# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import load_workbook, Workbook # 파일 불러오기
from openpyxl.styles import Alignment
from tkinter import messagebox as msgbox
import pandas as pd
import os

# 필요 파일 Open
CC_424 = load_workbook("424W_CC_2023-01-06.XLSX")
CC_424_ws = CC_424.active
WBS_424 = load_workbook("424W_WBS_2023-01-06.XLSX")
WBS_424_ws = WBS_424.active
CC_465 = load_workbook("465W_CC_2023-01-17.XLSX")
CC_465_ws = CC_465.active
Ref = load_workbook("Reference.xlsx")
Ref_WBS = Ref['WBS']
Ref_CC_465_ws = Ref['465W_CC List']
Ref_CC_424_ws = Ref['424W_CC List']

# Merge 파일 생성
wb = Workbook()
ws = wb.active
ws.title = 'Expense_Actual_Update'
# Merge 파일 Title 지정
ws['A1'] = 'Cost Center';ws['B1'] = 'Cost Element';ws['C1'] = 'Val/COArea Crcy';ws['D1'] = 'Partner object';ws['E1'] = 'Name';ws['F1'] = 'Org Cost Center';ws['G1'] = 'Partner object type'

##### 424W WBS 값 변환 및 Merge

# WBS 424의 Profit Center 값 반복
for cell in WBS_424_ws["A"]:
    # title 제외
    if cell.row == 1:
        continue
    # PC가 60040이면
    if cell.value == '60040':
        # WBS element의 앞에서 11까지만 참조
        WBS_element = WBS_424_ws["D"+cell.coordinate[1:]].value[:11]
        # WBS 참조표 값 반복
        for wbs in Ref_WBS['A']:
             #title 제외
            if wbs.row == 1:
                continue
            # 참조표의 wbs code와 update 파일의 wbs code가 같으면
            if wbs.value == WBS_element:
                # 참조표의 CC 값 저장
                CC = Ref_WBS["B"+wbs.coordinate[1:]].value
                ws["A"+cell.coordinate[1:]] = CC
    # PC가 60040이 아니면 그대로 PC 가져오기
    else:
        ws["A"+cell.coordinate[1:]] = cell.value
    # Org Cost Center에 Profit Center 복붙
    ws["F"+cell.coordinate[1:]] = cell.value
    # Account Number가 61790000이면 61790000W로 변환
    if WBS_424_ws["B"+cell.coordinate[1:]].value == '61790000':
        ws["B"+cell.coordinate[1:]].value = '61790000W'
    # Account Number가 61790000가 아니면 그대로
    else:
        ws["B"+cell.coordinate[1:]].value = WBS_424_ws["B"+cell.coordinate[1:]].value

for cols in WBS_424_ws["C:D"]:
    for cell in cols:
        if cell.row == 1:
            continue
        elif cell.column == 3: # In profit center local currency -> Val/COArea Crcy
            ws["C"+cell.coordinate[1:]] = cell.value
            ws["C"+cell.coordinate[1:]].number_format = openpyxl.styles.numbers.builtin_format_code(3)  # '#,##0'
        elif cell.column == 4: # WBS element -> Partner object
            ws["D"+cell.coordinate[1:]] = cell.value

# 마지막 행 값에 추가해서 데이터를 입력하기 위한 인덱스 정의
i = ws.max_row
j = i;k = i;l = i;m = i;o = i

##### 424W CC 값 변환 및 Merge
for cell_CC in CC_424_ws["A"]:
    # title 제외
    if cell_CC.row == 1:
        continue
    # WBS 참조표 값 반복
    for CC in Ref_CC_424_ws['A']:
        #title 제외
        if CC.row == 1:
            continue
        # type 통일
        CC.value = str(CC.value)
        # 참조표의 PCCC와 update 파일의 Cost Center가 같으면
        if cell_CC.value == CC.value:
            # 참조표의 CC 값 저장
            CC_final = Ref_CC_424_ws["B"+CC.coordinate[1:]].value
            ws.cell(row = i, column=1).value = CC_final  #A열(Column=1)에 행을 바꾸면서 입력
            i+=1  #행 값 증가
for cols in CC_424_ws["B:F"]:
    for cell in cols:
        if cell.row == 1:
            continue
        if cell.column == 2: # Cost Element
            ws.cell(row = j, column=2).value = cell.value
            j+=1
        elif cell.column == 3: # Val/COArea Crcy
            ws.cell(row = k, column=3).value = cell.value
            ws.cell(row = k, column=3).number_format = openpyxl.styles.numbers.builtin_format_code(3)  # '#,##0'
            k+=1
        elif cell.column == 4: # Partner object
            ws.cell(row = l, column=4).value = cell.value
            l+=1
        elif cell.column == 5: # Partner object type
            ws.cell(row = m, column=7).value = cell.value
            m+=1
        else: # Name
            ws.cell(row = o, column=5).value = cell.value
            o+=1

# 마지막 행 값에 추가해서 데이터를 입력하기 위한 인덱스 정의
i = ws.max_row
j = i;k = i;l = i;m = i;

###### 465W CC 값 변환 및 Merge

for cell_CC in CC_465_ws["A"]:
    # title 제외
    if cell_CC.row == 1:
        continue
    # WBS 참조표 값 반복
    for CC in Ref_CC_465_ws['A']:
        #title 제외
        if CC.row == 1:
            continue
        # type 통일
        CC.value = str(CC.value)
        # 참조표의 AP0 465W와 update 파일의 Profit Center가 같으면
        if cell_CC.value == CC.value:
            # 참조표의 CC 값 저장
            CC_final = Ref_CC_465_ws["B"+CC.coordinate[1:]].value
            ws.cell(row = j, column=1).value = CC_final  #A열(Column=1)에 행을 바꾸면서 입력
            j+=1  #행 값 증가

for cols in CC_465_ws["A:C"]:
    for cell in cols:
        if cell.row == 1:
            continue
        if cell.column == 1: # Profit Center -> Org Cost Center
            ws.cell(row = k, column=6).value = cell.value
            k+=1
        elif cell.column == 2: # Account Number -> Cost Center
            ws.cell(row = l, column=2).value = cell.value
            l+=1
        elif cell.column == 3: # In profit center local currency -> Val/COArea Crcy
            ws.cell(row = m, column=3).value = cell.value
            ws.cell(row = m, column=3).number_format = openpyxl.styles.numbers.builtin_format_code(3)  # '#,##0'
            m+=1
# 마지막 총합 삭제
ws.delete_rows(ws.max_row)

######### 삭제 조건에 따라 행 삭제
ws['H1'] = 'Delete' # title

for cell in ws["B"]:
    if cell.row == 1:
        continue
    cell.value = str(cell.value)
    try:
        if cell.value == '91490000' or cell.value == '91980130':
            if 'CS' in ws["E"+cell.coordinate[1:]].value:
                ws["H" + cell.coordinate[1:]].value = 'Delete'
        if cell.value == '81210000' or cell.value == '81251100' or cell.value == '81230000':
            if 'SEG' in ws["E" + cell.coordinate[1:]].value:
                ws["H" + cell.coordinate[1:]].value = 'Delete'
        if ws["F"+cell.coordinate[1:]].value == '60020':
            ws["H" + cell.coordinate[1:]].value = 'Delete'
        if 'SC' in ws["D" + cell.coordinate[1:]].value:
            ws["H" + cell.coordinate[1:]].value = 'Delete'
    except:
        continue
        
##### 셀 너비 조정 및 단위 변경
for column_cells in ws.columns:
    length = max(len(str(cell.value))*1.1 for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length
    
# 필터 생성
ws.auto_filter.ref = ws.dimensions # 필터 범위 지정
ws.auto_filter.add_filter_column(0, []) # 1행에 필터 생성하기
        
        
######## 최종 파일 생성
final = wb.copy_worksheet(ws)
final.title = 'Final'

######## Final 값들만 남기고 Delete되는 행 데이터 지우기
for row in final[2:final.max_row]: # 전체 행
    for cell in row:
        if final["H" + cell.coordinate[1:]].value == 'Delete': # H열이 Delete면
            cell.value = "" # 데이터 지우기
wb.save('Expense_Actual_Update.xlsx')

# # Delete 값에 해당하는 데이터 삭제
df = pd.read_excel('Expense_Actual_Update.xlsx', sheet_name='Final')
df = df.dropna(how='all')
df.to_excel('Expense_Actual_Update_Final.xlsx', sheet_name = 'Final', index=False)

wb.remove(final)

# final_wb = load_workbook("Expense_Actual_Update_Final.xlsx")
# final = final_wb.active

    
# final_wb.save('Expense_Actual_Update_Final.xlsx')
# final_wb.close()
######## 최종 파일 생성

final_wb = load_workbook("Expense_Actual_Update_Final.xlsx")
final = final_wb.active
Final = wb.create_sheet('Final')
for row in final:
    for cell in row:
        Final[cell.coordinate].value = cell.value

for cell in Final["C"]:
    cell.number_format = openpyxl.styles.numbers.builtin_format_code(3)

##### 셀 너비 조정 및 단위 변경
for column_cells in Final.columns:
    length = max(len(str(cell.value))*1.1 for cell in column_cells)
    Final.column_dimensions[column_cells[0].column_letter].width = length

# msgbox.showinfo("Complete", "파일 Merge가 완료되었습니다.")

wb.save('Expense_Actual_Update.xlsx')
wb.close()

#### 필요없는 파일 삭제
os.remove("Expense_Actual_Update_Final.xlsx")

print("Expense Actual Update has been done")