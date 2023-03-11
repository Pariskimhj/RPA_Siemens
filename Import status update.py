import requests
import json
from bs4 import BeautifulSoup as bs
from datetime import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import messagebox as msgbox


### 이 값은 계속 교체 해주어야할수도 있음
## 쿠키값은 개발자도구에서 Application -> cookies 
JSESSIONID = '00024iuvLsKw6hp1ESkp-iNAHqPgkcaFKjJNcUfA4brJblXfjRCeh4Wj7cY3dLNNWr_kvJRYO_Ppg1OSLw6YW9OlUvH0Ht6RdIcah6aPa8804EFZ4j187d5JpOTV4P5-AdAb:csp41'
### 개발자도구 network에서 내가하고 싶은 행동을 하면 뭐가 찍힘
def scrapUnipass(hblNoInfos):
    result = list()
    for (hblNo, year) in hblNoInfos:
        # Headers에서 발견
        tokenUrl = 'https://unipass.customs.go.kr/csp/myc/mainmt/MainMtCtr/menuExec.do'
        cookies = {
            'JSESSIONID': JSESSIONID
        }
        # Playload에서 발견
        tokenData = {
            'selectedId': 'MYC_MNU_00000450',
            'mblNo': '',
            'hblNo': hblNo,
            'blYy': year
        }
        # Post 방식으로 Request, 
        page = requests.post(tokenUrl, cookies=cookies, data=tokenData, verify=False)
        
        soup = bs(page.text, "html.parser")
        # 검색할 때 마다 변하는 고유한 값을 불러오기 위해, name은 고정
        savedToken = soup.find('input', {'name': 'MYC0405101Q_F2_savedToken'}).get('value')
        
        ############# 실제 데이터 뽑기 ##############
        prgsSttsUrl = 'https://unipass.customs.go.kr/csp/myc/bsopspptinfo/cscllgstinfo/ImpCargPrgsInfoMtCtr/retrieveImpCargPrgsInfoLst.do?savedToken=MYC0405101Q_F2_savedToken&MYC0405101Q_F2_savedToken=%s' % savedToken

        # 네트워크 -> Preview를 보고 내가 원하는 값을 가진 메뉴를 택한다
        # form data를 보면 어떻게 data를 요청해야 받을 수 있는지 나옴. 복붙하고 내가 원하는 값만 변수화
        ################ 진행상태  뽑아오기
        prgsSttsData = {
            'firstIndex': '0',
            'page': '1',
            'pageIndex': '1',
            'pageSize': '10',
            'pageUnit': '10',
            'recordCountPerPage': '10',
            'qryTp': '2',
            'cargMtNo': '',
            'mblNo': '',
            'hblNo': hblNo,
            'blYy': year
        }
        # 다시 Headers에 있는 Url에 요청함.
        prgsSttsResponse = requests.post(prgsSttsUrl, cookies=cookies, data=prgsSttsData, verify=False)
        # 요청 받은 결과 json형식으로 저장
        prgsSttsResponseJson = json.loads(prgsSttsResponse.text)
        # preview에서 resultlist에 원하는 값 불러오기

        # 진행상태
        prgsStts = prgsSttsResponseJson['resultList'][0]['prgsStts']
        # 통관진행상태
        prcsStcd = prgsSttsResponseJson['resultList'][0]['prcsStcd']
        
        result.append({
            '운송장번호': hblNo,
            '진행상태': prgsStts,
            '통관진행상태': prcsStcd
        })
    return result


# 엑셀에서 BL, year 끌고 오기
today = date.today().strftime("(%y.%m.%d)")
for i in range(2, 6):

    i = str(i)
    wb = load_workbook('수입현황_'+i+today+'.xlsx') # 수입 파일 열기
    ws = wb.active

    ws['U1'] = '진행상태'
    ws['T1'] = '통관진행상태'
    hblNoInfos = []
    col_B = ws["B"]

    for cell in col_B:
        # title 제외
        if cell.row == 1:
            continue
        if ws["O"+cell.coordinate[1:]].value == False or ws["P"+cell.coordinate[1:]].value == False: # 통관완료 또는 반출완료가 False이면
            if ws["L"+cell.coordinate[1:]].value < datetime.today(): # 입항일자가 오늘 이전이고
                BL = str(cell.value)
                Year = ws["L"+cell.coordinate[1:]].value.year
                hblNoInfos.append((BL, Year))
            else:
               cell.fill = PatternFill(fgColor="ffff00", fill_type='solid')
    # 데이터 가져오기
    Status = scrapUnipass(hblNoInfos)
    # 엑셀에 결과값 반영
    for s in Status:
        BL = s['운송장번호']
        진행상태 = s['진행상태']
        통관진행상태 = s['통관진행상태']
        for cell in col_B:
            if cell.value == BL:
                ws["U"+cell.coordinate[1:]].value = 진행상태
                ws["T"+cell.coordinate[1:]].value = 통관진행상태
                if 진행상태 == '반출완료':
                    ws["P"+cell.coordinate[1:]].value = True
                    ws["P"+cell.coordinate[1:]].fill = PatternFill(fgColor="ffff00", fill_type='solid')
                    cell.fill = PatternFill(fgColor="0066ff", fill_type='solid')
                elif 통관진행상태 == '수입신고수리' or 통관진행상태 == '반출신고':
                    ws["O"+cell.coordinate[1:]].value = True
                    ws["O"+cell.coordinate[1:]].fill = PatternFill(fgColor="ffff00", fill_type='solid')
                    cell.fill = PatternFill(fgColor="0066ff", fill_type='solid')
wb.save('수입현황_'+i+today+'.xlsx')
wb.close
msgbox.showinfo("Complete", "Import status updated.")
